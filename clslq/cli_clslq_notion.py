# -*- encoding: utf-8 -*-
'''cli_clslq_notion

Notion client in python, notion-py is required

Usage: clslq notion [OPTIONS]

[openpyxl](https://openpyxl-chinese-docs.readthedocs.io/zh_CN/latest/tutorial.html)

[pandas](https://www.pypandas.cn/)

Notion template on https://airy-skiff-4d0.notion.site/04143158def3413fb58e7dae4b2d9aff

'''


import click
import datetime
import string
import time
import json
import os
import re
import traceback
import pandas
# Support email
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill

from .clslq_config import ClslqConfigUnique
from .clslq_log import ClslqLogger
from notion_client import Client

from .templates import weekreport

clslog = ClslqLogger().log


class Report(object):
    """Base class of Report

    Args:
        object (wbname): Workbook name
    """

    def __init__(self, wbname):
        self.wb = Workbook()
        self.sht = self.wb.active
        self.default_border = Border(left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin',
                                                color='000000'),
                                     top=Side(border_style='thin',
                                              color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))
        self.wbname = wbname

    def render_html(self, clsconfig, title, database):
        """Render html form template

        Note that some of the email display methods only support inline-css style,
        This method support inline-css for table headers and rows.

        Args:
            title (str): Title of html and email subject
            df (object): Pandas DataFrame object
        """
        clslog.info("Render[Inline-CSS] html for {}".format(title))
        with open(self.wbname+'.html', encoding='utf-8', mode='w') as f:
            table = str('')
            summary = str('')
            plan = str('')
            i = 0
            j = 0
            for node in database['results']:
                item = node['properties']

                task_paln_template = """
                    <tr height="19" style="height:14.0pt;background: {color}">
                        <td style="border: 0.5pt solid #cfcfcf; vertical-align: middle; text-align: center;">{type}</td>
                        <td style="border: 0.5pt solid #cfcfcf;">{title}</td>
                        <td style="border: 0.5pt solid #cfcfcf; vertical-align: middle; text-align: center;">{state}</td>
                        <td style="border: 0.5pt solid #cfcfcf;">{problem}</td>
                        <td style="border: 0.5pt solid #cfcfcf;">{solve}</td>
                    </tr>
                """
                summary_template = """
                    <tr>
                        <td style="padding: 10px; background-color: rgba(204, 204, 204, 0.1)">
                        <span style="font-size: 16px; color: #81e4c3">●</span>&nbsp;
                        <span>
                            <span style="border-bottom: 1px dashed rgb(204, 204, 204); position: relative;">{summary}</span>
                        </span>
                        </td>
                    </tr>
                """
                def bgcolor(i): return '#F7F7F7' if i % 2 == 0 else '#fff'
                if self.content_parse_type(item, 0) == u'工作计划':
                    plan += task_paln_template.format(**{
                        'type': self.content_parse_type(item, 0),
                        'title': self.content_parse_title(item),
                        'state': self.content_parse_state(item, 0),
                        'problem': self.content_parse_richtext(item, u'问题'),
                        'solve': self.content_parse_richtext(item, u'解决方法'),
                        'color': bgcolor(i)
                    })
                    i = i + 1
                else:
                    table += task_paln_template.format(**{
                        'type': self.content_parse_type(item, 0),
                        'title': self.content_parse_title(item),
                        'state': self.content_parse_state(item, 0),
                        'problem': self.content_parse_richtext(item, u'问题'),
                        'solve': self.content_parse_richtext(item, u'解决方法'),
                        'color': bgcolor(j)
                    })
                    j = j + 1
                summarize_item = self.content_parse_richtext(item, u'评审、复盘、总结')
                if len(summarize_item):
                    summary += summary_template.format(**{
                        'summary': summarize_item
                    })
            html = string.Template(weekreport.wr_template)

            f.write(html.safe_substitute({
                "title": title,
                "table": table,
                "plan": plan,
                "user": clsconfig.get('user'),
                "department": clsconfig.get('department'),
                "summarize": summary
            }))

    def render_html_without_inline_css(self, title, df):
        """Render html form template, use pandas

        Note that some of the email display methods only support inline-css style,
        This method does not use any inline-css for table headers and rows.

        Args:
            title (str): Title of html and email subject
            df (object): Pandas DataFrame object
        """
        clslog.info("Render html for {}".format(title))
        with open(self.wbname+'.html', encoding='utf-8', mode='w') as f:

            task = df[df[u'分类'] != u'工作计划']
            plan = df[df[u'分类'] == u'工作计划']

            t = string.Template(weekreport.wr_template)
            f.write(t.safe_substitute({
                "title": title,
                "table": task.to_html(classes='tablestyle', index=False, na_rep=""),
                "plan": plan.to_html(classes='tablestyle', index=False, na_rep="")
            }))

    def send_email(self, config, title):
        """Send report email to receivers defined in .clslq.json

        Args:
            config (dict): Loaded json object from .clslq.json
            title (str): Unicode string as email subject
        """
        email = config.get('email')
        smtpserver = email['sender']['smtpserver']
        user = email['sender']['user']
        pwd = email['sender']['pwd']
        receivers = email['receivers']
        clslog.info("{} {} {}".format(smtpserver, user, title))

        msg = MIMEMultipart()
        msg['From'] = "{}".format(user)
        msg['To'] = ",".join(receivers)
        msg['Subject'] = Header(title, 'utf-8')
        with open(self.wbname+'.html', "r", encoding='utf-8') as f:
            msg.attach(MIMEText(f.read(), 'html', 'utf-8'))
        try:
            smtp = smtplib.SMTP()
            smtp.connect(smtpserver)
            smtp.login(user, pwd)
            smtp.sendmail(user, receivers, msg.as_string())
            clslog.info('Email sent to {} done'.format(receivers))
            smtp.quit()
            smtp.close()
        except Exception as e:
            clslog.critical(e)

    def remove_files(self):
        """Removes all generated files"""
        if os.path.exists(self.wbname+'.html'):
            os.remove(self.wbname+'.html')
        if os.path.exists(self.wbname+'.xlsx'):
            os.remove(self.wbname+'.xlsx')


class WeekReport(Report):
    def set_worksheet_head(self, head):
        """Set sheet head

        Args:
            head (worksheet head): string
        """
        self.sht.merge_cells('A1:F1')

        self.sht['A1'] = u"本周工作情况"+head.replace('-', "")
        self.sht['A1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sht['A1'].font = Font(
            color=colors.BLACK, b=True, size=14)
        self.sht['A1'].fill = PatternFill("solid", fgColor="00FF8080")
        self.sht.row_dimensions[1].height = 20

    def set_worksheet_title(self):
        u"""openpyxl 不支持按列设置样式
        """
        self.sht['A2'] = u"分类"
        self.sht['B2'] = u"事项"
        self.sht['C2'] = u"进展"
        self.sht['D2'] = u"问题"
        self.sht['E2'] = u"解决"
        self.sht['F2'] = u"评审、复盘、总结"
        for col in range(1, 7):
            self.sht.cell(column=col, row=2).font = Font(
                name=u'微软雅黑', bold=True, size=12)

    def content_parse_title(self, item):
        result = None
        try:
            title = item[u'名称']['title']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_state(self, item, row):
        result = ''
        try:
            for i in item[u'状态']['multi_select']:
                if i['name'] == u'进行中':
                    for col in range(1, 7):
                        self.sht.cell(column=col, row=row).fill = PatternFill(
                            "solid", fgColor="00CCFFCC")
                if i['name'] == u'遇问题':
                    for col in range(1, 7):
                        self.sht.cell(column=col, row=row).fill = PatternFill(
                            "solid", fgColor="00FFCC99")
                result = "{} {}".format(result, i['name'])
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_type(self, item, row):
        result = ''
        try:
            result = item[u'分类']['select']['name']
            if result == u'工作计划':
                self.sht.cell(column=6, row=row).value = u"下周工作计划"
                for col in range(1, 7):
                    self.sht.cell(column=col, row=row).fill = PatternFill(
                        "solid", fgColor="00CCFFCC")
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_richtext(self, item, text):
        """Parse Notion Column content

        Args:
            item (dict): Notion Column content
            text (str): Unicode string means column title

        Returns:
            str: Cell result
        """
        result = ''
        try:
            title = item[text]['rich_text']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result.replace('\n', ' ')

    def excel_worksheet_fill(self, item, row):
        self.sht['A'+str(row)] = self.content_parse_type(item, row)
        self.sht['B'+str(row)] = self.content_parse_title(item)
        self.sht['C'+str(row)] = self.content_parse_state(item, row)
        self.sht['D'+str(row)] = self.content_parse_richtext(item, u'问题')
        self.sht['E'+str(row)] = self.content_parse_richtext(item, u'解决方法')
        self.sht['F'+str(row)] = self.content_parse_richtext(item, u'评审、复盘、总结')

        for c in ('A', 'C'):
            self.sht.column_dimensions[c].width = 10
            for cell in self.sht[c]:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.default_border
        for c in ('B', 'F', 'D', 'E'):
            self.sht.column_dimensions[c].width = 30
            for cell in self.sht[c]:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.default_border

    def excel_worksheet_dump(self, wbname, database):
        # Change sheet name
        self.sht.title = 'WR'
        self.set_worksheet_head(wbname)
        self.set_worksheet_title()

        row = 3
        for node in database['results']:
            item = node['properties']
            self.excel_worksheet_fill(item, row)
            row = row + 1
        self.wb.save(filename=wbname+'.xlsx')
        self.wb.close()

    def pandas_df_fill(self, database):
        _type = []
        _title = []
        _state = []
        _problem = []
        _solve = []
        _conclusion = []
        for node in database['results']:
            item = node['properties']
            _type.append(self.content_parse_type(item, 0))
            _title.append(self.content_parse_title(item))
            _state.append(self.content_parse_state(item, 0))
            _problem.append(self.content_parse_richtext(item, u'问题'))
            _solve.append(self.content_parse_richtext(item, u'解决方法'))
            _conclusion.append(self.content_parse_richtext(item, u'评审、复盘、总结'))
        data = {
            u'分类': pandas.Series(_type, index=range(len(_type))),
            u'事项': pandas.Series(_title, index=range(len(_title))),
            u'进展': pandas.Series(_state, index=range(len(_state))),
            u'问题': pandas.Series(_problem, index=range(len(_problem))),
            u'解决': pandas.Series(_solve, index=range(len(_solve))),
            u'评审复盘总结备注': pandas.Series(_conclusion, index=range(len(_conclusion)))
        }
        self.df = pandas.DataFrame(data)
        return self.df


def cli_week(client, clsconfig, excel, remove, force, send):
    for i in client.search()['results']:
        if i['object'] == 'database':
            try:

                database = client.databases.query(i['id'])
                title = i['title']

                for t in title:
                    plain_text = t['plain_text'].strip().replace(
                        ' → ', '~')
                    plain_text_head = t['plain_text'][0:3]
                    value = re.compile(r'^[0-9]+[0-9]$')
                    # Handle valid report database only
                    if plain_text_head == 'WRT':
                        break
                    if value.match(plain_text_head) == None:
                        break
                    enddate = datetime.datetime.fromisoformat(
                        t['mention']['date']['end'])
                    nowdate = datetime.datetime.now()
                    # Use BT-Panel timer task to trigger
                    if not force:
                        if nowdate.weekday() != 5:  # 0~6 means Monday~Sunday
                            clslog.warning("Today is not Saturday")
                            break
                    if abs(enddate-nowdate) > datetime.timedelta(days=5):
                        click.secho("End:{} now:{} weekday:{} delta:{} Week report expired".format(
                            enddate, nowdate, nowdate.weekday(), abs(enddate-nowdate)), fg='green')
                        break
                    wrp = WeekReport(plain_text)
                    wtitle = "{}({})".format(
                        clsconfig.get('wr_title_prefix'), plain_text)

                    if excel:
                        wrp.excel_worksheet_dump(plain_text, database)
                        df = pandas.read_excel(
                            plain_text+'.xlsx', sheet_name='WR', header=1)
                    else:
                        df = wrp.pandas_df_fill(database)
                    pandas.set_option('colheader_justify', 'center')
                    df.style.hide_index()  # Hide index col

                    wrp.render_html(clsconfig, wtitle, database)
                    if send:
                        wrp.send_email(clsconfig, wtitle)
                    if remove:
                        wrp.remove_files()

            except Exception as e:
                clslog.error(e)
                traceback.print_exc(e)


def cli_month(client, clsconfig, remove, force):

    nowdate = datetime.datetime.now()
    mtitle = "{}({}{:02})".format(
        clsconfig.get('mr_title_prefix'), nowdate.year, nowdate.month)
    click.secho("{}".format(mtitle), fg='blue')


@click.option('--rtype',
              '-t',
              required=True,
              default='week',
              type=click.Choice(['week', 'month', 'all']),
              help='Choose a type to generate report')
@click.option('--excel',
              '-e',
              flag_value='GenerateExcel',
              default=False,
              help='Generate xlsx excel file or not')
@click.option('--remove',
              '-r',
              flag_value='RemoveFiles',
              help='Remove files or not')
@click.option('--force',
              '-f',
              flag_value='force',
              default=False,
              help='Force generate right now, otherwise limited by valid datetime')
@click.option('--send',
              '-s',
              flag_value='send',
              default=False,
              help='Send email or not')
@click.option('--config',
              '-c',
              type=click.Path(exists=True),
              default='.clslq.json',
              help='CLSLQ config use {} as default.'.format('.clslq.json'))
@click.command(context_settings=dict(
    allow_extra_args=True,
    ignore_unknown_options=True,
),
    help="Notion Report Generator.")
def notion(rtype, config, excel, remove, force, send):

    clsconfig = ClslqConfigUnique(file=config)
    if clsconfig.get('secrets_from') is None:
        click.secho(
            "Make sure notion secret code is valid in .clslq.json", fg='red')
        return

    client = Client(auth=clsconfig.get('secrets_from'),
                    notion_version="2021-08-16")

    click.secho("Report type: {} Notion Accounts:".format(rtype), fg='yellow')
    for u in client.users.list()['results']:
        click.secho("{:8s} {}".format(u['name'], u['id']), fg='green')
    if 'week' == rtype:
        click.secho("Week report generator", fg='green')
        cli_week(client, clsconfig, excel, remove, force, send)
    elif 'month' == rtype:
        click.secho("Month report generator", fg='green')
        cli_month(client, clsconfig, remove, force, send)
    else:
        click.secho("All reports generator", fg='green')
        cli_week(client, clsconfig, excel, remove, force, send)
        cli_month(client, clsconfig, remove, force, send)
